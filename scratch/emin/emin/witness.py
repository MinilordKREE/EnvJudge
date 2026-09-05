"""W probe (zero LLM cost) over all verified_400 records.

Per task:
  parse_ok        answer_position parses on both paths (substrate range_boundaries, official _generate_cell_names)
  n_cells         cells in answer_position
  h               cells in answer_position that read None from the golden under data_only=True WITHOUT recalc
  n_formula       cells in answer_position whose golden stores a formula (data_only=False value starts with "=")
  recalc_ok       LibreOffice recalc of a golden copy succeeded
  W_a_pass        official path golden-vs-golden (two independent recalculated copies) passes compare_workbooks
  W_pass          W_a_pass and parse_ok  (W fails  ->  L0 candidate)
  sub_gvg         substrate path golden-vs-golden verdict (trivially pass unless parse fails)
  sub_recalc_vs_raw  substrate verdict of recalc(golden) against raw golden: FAIL means a correct,
                     fully-computed output would be rejected by the substrate path (h>0 mechanism)
Output: results/emin/witness_probe.csv (append-only deliverable; rerun writes a new timestamped copy if it exists).
"""

from __future__ import annotations

import csv
import shutil
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import range_boundaries

from emin.contracts import recalc
from emin.data import all_records, load_task, split_of
from emin.score import compare_official, score_substrate
from emin.settings import config
from emin.third_party.online_judge_eval import _generate_cell_names

FIELDS = ["task_id", "split", "instruction_type", "answer_position", "parse_ok", "parse_detail", "n_cells", "h", "n_formula",
          "recalc_ok", "recalc_detail", "W_a_pass", "W_a_reason", "W_pass", "sub_gvg", "sub_recalc_vs_raw", "sub_recalc_vs_raw_reason",
          "load_error"]


def _segments(answer_position: str):
    for raw in str(answer_position).split(","):
        part = raw.strip()
        if not part:
            continue
        if "!" in part:
            sheet, rng = part.split("!", 1)
        else:
            sheet, rng = None, part
        yield (sheet.strip().strip("'\"") if sheet else None), rng.strip().strip("'\"")


def probe_record(record: dict, soffice: str, timeout: int) -> dict:
    row = {k: "" for k in FIELDS}
    task_id = str(record["id"])
    row.update(task_id=task_id, split=split_of(task_id), instruction_type=record.get("instruction_type", ""))
    try:
        task = load_task(record)
    except Exception as exc:  # missing files etc.
        row.update(answer_position=record.get("answer_position", ""), load_error=f"{type(exc).__name__}: {exc}"[:200], parse_ok=False, W_pass=False)
        return row
    ap = task.answer_position
    row["answer_position"] = ap
    # (c) parse
    parse_ok, detail, n_cells = True, "", 0
    try:
        for sheet, rng in _segments(ap):
            c1, r1, c2, r2 = range_boundaries(rng)
            n_cells += (c2 - c1 + 1) * (r2 - r1 + 1)
            _generate_cell_names(rng)
    except Exception as exc:
        parse_ok, detail = False, f"{type(exc).__name__}: {exc}"[:120]
    row.update(parse_ok=parse_ok, parse_detail=detail, n_cells=n_cells)
    # (b) hidden None and formula counts on the raw golden
    h = n_formula = 0
    try:
        wb_v = openpyxl.load_workbook(task.golden_path, data_only=True, read_only=True)
        wb_f = openpyxl.load_workbook(task.golden_path, data_only=False, read_only=True)
        try:
            for sheet, rng in _segments(ap):
                sname = sheet or wb_v.sheetnames[0]
                if sname not in wb_v.sheetnames:
                    continue
                c1, r1, c2, r2 = range_boundaries(rng)
                for r_v, r_f in zip(
                    wb_v[sname].iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2, values_only=True),
                    wb_f[sname].iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2, values_only=True),
                ):
                    for v, f in zip(r_v, r_f):
                        if v is None:
                            h += 1
                        if isinstance(f, str) and f.startswith("="):
                            n_formula += 1
        finally:
            wb_v.close()
            wb_f.close()
    except Exception as exc:
        row["load_error"] = f"{type(exc).__name__}: {exc}"[:200]
    row.update(h=h, n_formula=n_formula)
    # substrate golden-vs-golden
    s = score_substrate(task.golden_path, task.golden_path, ap)
    row["sub_gvg"] = s.reason
    # (a) official path golden-vs-golden with two independent recalculated copies
    with tempfile.TemporaryDirectory(prefix="emin_w_") as tmp:
        a = Path(tmp) / "a.xlsx"
        b = Path(tmp) / "b.xlsx"
        shutil.copyfile(task.golden_path, a)
        shutil.copyfile(task.golden_path, b)
        ok_a, d_a = recalc(str(a), soffice, timeout)
        ok_b, d_b = recalc(str(b), soffice, timeout) if ok_a else (False, "skipped")
        row.update(recalc_ok=ok_a and ok_b, recalc_detail=(d_a if not ok_a else d_b)[:120])
        if ok_a and ok_b:
            sc = compare_official(a, b, ap, task.instruction_type)
            row.update(W_a_pass=sc.passed, W_a_reason=sc.reason)
            s2 = score_substrate(a, task.golden_path, ap)   # recalculated golden judged by the substrate path against the raw golden
            row.update(sub_recalc_vs_raw=s2.verdict, sub_recalc_vs_raw_reason=s2.reason)
        else:
            row.update(W_a_pass=False, W_a_reason="recalc_error")
    row["W_pass"] = bool(row["W_a_pass"]) and parse_ok
    return row


def main() -> None:
    cfg = config()
    out_dir = Path(cfg["results_dir"])
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "witness_probe.csv"
    if out.exists():
        out = out_dir / f"witness_probe_{time.strftime('%Y%m%dT%H%M%SZ', time.gmtime())}.csv"
    records = all_records()
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=cfg["concurrency"]["recalc"]) as pool:
        rows = list(pool.map(lambda r: probe_record(r, cfg["soffice"], cfg["execution"]["recalc_timeout_seconds"]), records))
    rows.sort(key=lambda r: (r["split"], r["task_id"]))
    with open(out, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)
    pool_rows = [r for r in rows if r["split"] == "pool"]
    print(f"wrote {out} ({len(rows)} rows) in {time.time() - t0:.0f}s")
    print("all: W_pass", sum(bool(r["W_pass"]) for r in rows), "h>0", sum(int(r["h"] or 0) > 0 for r in rows),
          "recalc_ok", sum(bool(r["recalc_ok"]) for r in rows), "load_error", sum(bool(r["load_error"]) for r in rows))
    print("pool 76: W_pass", sum(bool(r["W_pass"]) for r in pool_rows), "h>0", sum(int(r["h"] or 0) > 0 for r in pool_rows),
          "sub_recalc_vs_raw FAIL", sum(r["sub_recalc_vs_raw"] == "FAIL" for r in pool_rows))


if __name__ == "__main__":
    main()
