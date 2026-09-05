"""Every reason string on both scoring paths, plus the c1/c2 contracts."""

from __future__ import annotations

import ast
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from emin import contracts
from emin.score import REASONS, compare_official, score_official, score_substrate
from emin.settings import config


def wb(path: Path, values: dict[str, object], sheet: str = "S", formulas: dict[str, str] | None = None) -> Path:
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = sheet
    for k, v in values.items():
        ws[k] = v
    for k, f in (formulas or {}).items():
        ws[k] = f
    book.save(path)
    return path


@pytest.fixture
def golden(tmp_path):
    return wb(tmp_path / "g.xlsx", {"A1": 1, "A2": 2.0, "B1": "x"})


def test_substrate_pass(tmp_path, golden):
    out = wb(tmp_path / "o.xlsx", {"A1": 1, "A2": 2, "B1": "x"})
    assert score_substrate(out, golden, "S!A1:B2").reason == "pass"


def test_substrate_value_mismatch(tmp_path, golden):
    out = wb(tmp_path / "o.xlsx", {"A1": 1, "A2": 3, "B1": "x"})
    s = score_substrate(out, golden, "S!A1:B2")
    assert (s.verdict, s.reason) == ("FAIL", "value_mismatch")


def test_substrate_sheet_missing(tmp_path, golden):
    out = wb(tmp_path / "o.xlsx", {"A1": 1}, sheet="Other")
    assert score_substrate(out, golden, "S!A1:B2").reason == "sheet_missing"


def test_substrate_workbook_missing(tmp_path, golden):
    assert score_substrate(tmp_path / "nope.xlsx", golden, "S!A1").reason == "workbook_missing"


def test_substrate_workbook_parse_error(tmp_path, golden):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not a zip")
    assert score_substrate(bad, golden, "S!A1").reason == "workbook_parse_error"


def test_substrate_answer_position_parse_error(tmp_path, golden):
    out = wb(tmp_path / "o.xlsx", {"A1": 1})
    assert score_substrate(out, golden, "").reason == "answer_position_parse_error"
    assert score_substrate(out, golden, "S!garbage").reason == "answer_position_parse_error"


def test_substrate_formula_reads_none(tmp_path):
    """The substrate mechanism under test: a formula cell without cached value reads None."""
    g = wb(tmp_path / "g.xlsx", {"A1": 1, "A2": 2}, formulas={"A3": "=A1+A2"})
    o = wb(tmp_path / "o.xlsx", {"A1": 1, "A2": 2, "A3": 3})
    assert score_substrate(o, g, "S!A3").reason == "value_mismatch"   # golden None vs 3
    assert score_substrate(g, g, "S!A3").reason == "pass"              # None vs None


def test_official_reasons(tmp_path, golden):
    assert compare_official(golden, wb(tmp_path / "o1.xlsx", {"A1": 1, "A2": 2, "B1": "x"}), "S!A1:B2").reason == "pass"
    assert compare_official(golden, wb(tmp_path / "o2.xlsx", {"A1": 1, "A2": 9, "B1": "x"}), "S!A1:B2").reason == "value_mismatch"
    assert compare_official(golden, wb(tmp_path / "o3.xlsx", {"A1": 1}, sheet="Z"), "S!A1:B2").reason == "sheet_missing"
    assert compare_official(golden, tmp_path / "nope.xlsx", "S!A1").reason == "workbook_missing"
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"zzz")
    assert compare_official(golden, bad, "S!A1").reason == "workbook_parse_error"
    assert compare_official(golden, wb(tmp_path / "o4.xlsx", {"A1": 1}), "S!A1:B2:C3").reason == "answer_position_parse_error"


def test_official_recalc_error(tmp_path, golden):
    out = wb(tmp_path / "o.xlsx", {"A1": 1})
    s = score_official(out, golden, "S!A1", "", tmp_path / "work", soffice="/nonexistent/soffice")
    assert s.reason == "recalc_error"
    assert score_official(tmp_path / "missing.xlsx", golden, "S!A1", "", tmp_path / "work").reason == "workbook_missing"


@pytest.mark.skipif(not Path(config()["soffice"]).is_file(), reason="LibreOffice not available")
def test_official_recalc_computes_formulas(tmp_path):
    g = wb(tmp_path / "g.xlsx", {"A1": 1, "A2": 2}, formulas={"A3": "=A1+A2"})
    o = wb(tmp_path / "o.xlsx", {"A1": 1, "A2": 2, "A3": 3})
    s = score_official(o, g, "S!A3", "", tmp_path / "work")
    assert s.reason == "pass", s
    assert score_substrate(o, g, "S!A3").reason == "value_mismatch"


def test_reason_vocabulary_closed():
    assert set(REASONS) == {"pass", "value_mismatch", "sheet_missing", "workbook_missing", "workbook_parse_error",
                            "recalc_error", "answer_position_parse_error"}


# ---- contracts -------------------------------------------------------------------------

def test_c1_variants():
    assert contracts.unwrap_fence("```python\nx = 1\n```") == "x = 1"
    assert contracts.unwrap_fence("Here:\n```python3\nx = 1\n```\nBye") == "x = 1"
    assert contracts.unwrap_fence("```\nx = 1\n```") == "x = 1"
    assert contracts.unwrap_fence("~~~py\nx = 1\n~~~") == "x = 1"
    assert contracts.unwrap_fence("```bash\nls\n```\n```python\nx = 2\n```") == "x = 2"
    assert contracts.unwrap_fence("```python\nx = 1\n```\n```python\nx = 2\n```") == "x = 1"   # first block
    assert contracts.unwrap_fence("x = 3") == "x = 3"
    assert contracts.unwrap_fence("Let's produce the code.```python\nx = 4\n```") == "x = 4"   # inline opening fence
    assert contracts.unwrap_fence("```python\nx = 5\n```trailing") == "x = 5"
    with pytest.raises(ValueError):
        contracts.unwrap_fence("")


def test_c2_wrapper_saves_on_exception(tmp_path):
    out = tmp_path / "out.xlsx"
    code = (
        "import openpyxl\n"
        "def main():\n"
        "    wb = openpyxl.Workbook()\n"
        "    wb.active['A1'] = 7\n"
        "    raise RuntimeError('boom before save')\n"
        "main()\n"
    )
    wrapped = contracts.run_with_save_on_exception(code)
    ast.parse(wrapped)
    script = f"OUTPUT_PATH = {str(out)!r}\n" + wrapped
    r = subprocess.run([sys.executable, "-c", script], capture_output=True, text=True)
    assert r.returncode != 0 and "boom before save" in r.stderr and contracts.C2_MARKER in r.stderr
    assert openpyxl.load_workbook(out).active["A1"].value == 7


def test_c2_wrapper_noop_on_success(tmp_path):
    out = tmp_path / "out.xlsx"
    code = "import openpyxl\nwb = openpyxl.Workbook()\nwb.active['A1'] = 1\nwb.save(OUTPUT_PATH)\n"
    script = f"OUTPUT_PATH = {str(out)!r}\n" + contracts.run_with_save_on_exception(code)
    r = subprocess.run([sys.executable, "-c", script], capture_output=True, text=True)
    assert r.returncode == 0 and contracts.C2_MARKER not in r.stderr


def test_c2_passes_substrate_isolation_validator():
    from rethinkskill_spreadsheetbench.runtime import validate_generated_code_isolation
    code = "from __future__ import annotations\nimport openpyxl\nwb = openpyxl.load_workbook(INPUT_PATH)\nwb.save(OUTPUT_PATH)\n"
    validate_generated_code_isolation(contracts.run_with_save_on_exception(code))
